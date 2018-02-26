import math
from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt

# opening .xlsx file
wb = load_workbook('mini-project1-data.xlsx')                                                                            
ws = wb[wb.sheetnames[0]]                                          
ws = wb.active                                              

# Global Variables
questionDistributions = {}
questionMeanVarEnt = {}

''' Parses each column using openpyxl, then computes the
    and returns the response distribution of the specific 
    question. '''
def distributionOfColumn(columnLabel):
    questionResponses = []
    responses = []

    # parse .xlsx file and add to list
    for row in range(2, ws.max_row + 1):
        for column in columnLabel:
            cellName = "{}{}".format(column, row)
            responses.append(ws[cellName].value)  

    # populate list with rates per response
    for i in range(5):
        questionResponses.append(responses.count(i + 1))
    
    return np.divide(questionResponses, len(responses))


''' Computes and returns the mean, variance, and entropy  
    for a specific question based on the response distributions. 
    Makes use of numpy for the list arithmetic. '''
def computeMeanVarEnt(response, probs):
    negProb = [-i for i in probs]
    expectation = sum(np.multiply(response, probs))
    doubleExpectation = sum(np.multiply(np.square(response), probs))
    variance = np.subtract(doubleExpectation, np.square(expectation))
    entropy = sum(np.multiply(negProb, np.log2(probs)))
    return expectation, variance, entropy


''' Computes the Hellinger distance of each possible question pair. 
    Returns a dictionary of the values with an accompanying key 
    describing the two questions.'''
def calculateHellinger(distributions, questionTitle):
    hellinDict = {}
    divSqRtTwo = 1 / math.sqrt(2)
    for qA in range(26):
        for qB in range(qA + 1, 26):
            probOfA = distributions.get(questionTitle[qA])
            probOfB = distributions.get(questionTitle[qB])
            temp = math.sqrt(sum(np.square(np.subtract(np.sqrt(probOfA), np.sqrt(probOfB)))))
            hellin = temp * divSqRtTwo

            # format key and create dictionary
            hellinKey = "%d, %d" % (qA + 1, qB + 1)
            hellinDict.update({hellinKey:hellin})
    return hellinDict


''' Draws the bar graph for all the distributions per response
    for each question.'''
def distributionBarGraph(distributions, questionTitle):
    colors = ['g', 'b', 'r', 'y', 'k']

    # bar graph setup
    plt.figure(figsize=(13, 4))
    index = np.arange(len(distributions.keys()))
    barWidth = 0.15

    # creates graph data
    for rating in range(5):
        counter = barWidth * (rating + 1)
        probabilityOfResponse = tuple([probs[rating] for probs in distributions.values()])
        plt.bar(index + counter, probabilityOfResponse, barWidth, 
                color=colors[rating], label=(rating + 1))   

    plt.xlabel('Questions')
    plt.ylabel('Distributions')
    plt.xticks(index + (barWidth * 3), tuple(questionTitle))
    plt.legend()
    plt.show()   


''' Draws the bar graph for the hellinger distance spreads for 
    each pair of question. '''
def hellingerBarGraph(distances):
    plt.figure(figsize=(10, 5))
    index = np.arange(len(distances.keys()))
    barWidth = 0.75

    plt.bar(index, tuple(distances.values()), barWidth, color='b')
    plt.xlabel('Question Pairs')
    plt.ylabel('Hellinger Distances')
    plt.tick_params(axis='x', which='both', bottom='off', labelbottom='off')
    plt.show()


''' Writes the dictionary data passed to a .txt file in a 
    readable format. '''
def writeDataToFile(filename, data, dictValueNum):
    title = ['mean', 'variance', 'entropy']
    f = open(filename, "w")
    for key, value in data.items():
        if dictValueNum <= 1:
            f.write("%s: %s\n" % (key, round(value, 4)))
        else:
            f.write("%s\n" % key)
            for i in range(dictValueNum):
                if dictValueNum == 3:
                    f.write("  %s: %s\n" % (title[i], round(value[i], 3)))
                else:
                    f.write("  %d: %s\n" % (i+1, round(value[i], 3)))  
    f.close()

# Main
answer = list(range(1, 6))

# loop that creates dictionaries to hold the contents of 
# the distributions, mean, variance, and entropy
for i in range(26):
    colLetter = chr(65 + i)
    questionName = "Q%d" % (i + 1)
    
    questionProb = distributionOfColumn(colLetter)
    questionMean, questionVar, questionEnt = computeMeanVarEnt(answer, questionProb)
 
    questionMeanVarEnt.update({questionName:[questionMean, questionVar, questionEnt]})          
    questionDistributions.update({questionName:questionProb})           

# calculates the hellinger distance for all possible values
keys = [key for key, values in questionDistributions.items()]


distances = calculateHellinger(questionDistributions, keys)
distanceMaxKey = max(distances, key=distances.get)
distanceMinKey = min(distances, key=distances.get)

# writes all the data to file for easy reading, also creates the 
# two graphs and prints the min and max hellinger pairs.
writeDataToFile("question_distributions.txt", questionDistributions, 5)
writeDataToFile("question_mean_var_ent.txt", questionMeanVarEnt, 3)
writeDataToFile("question_hellinger.txt", distances, 1)
distributionBarGraph(questionDistributions, keys)
hellingerBarGraph(distances)
print("%s: %s" % (distanceMaxKey, round(distances.get(distanceMaxKey), 3)))
print("%s: %s" % (distanceMinKey, round(distances.get(distanceMinKey), 3)))