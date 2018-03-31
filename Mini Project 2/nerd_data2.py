from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt
import math

wb = load_workbook('mini-project2-data-v2.xlsx')
ws = wb[wb.sheetnames[0]]
ws = wb.active

femaleQuestionDistributions = {}
maleQuestionDistributions = {}
hellingerDistances = {}

def distributionOfColumns(columnLabel):
    maleQuestionResponses = []
    femaleQuestionResponses = []
    maleResponses = []
    femaleResponses = []

    for row in range(2, ws.max_row + 1):
        for column in columnLabel:
            cellName = "{}{}".format(column, row)
            genderCell = "{}{}".format('AA', row)
            if ws[cellName].value > 0:
                if ws[genderCell].value == 1:
                    femaleResponses.append(ws[cellName].value)
                else:
                    maleResponses.append(ws[cellName].value)

    for i in range(5):
        femaleQuestionResponses.append(femaleResponses.count(i + 1))
        maleQuestionResponses.append(maleResponses.count(i + 1))
    
    femaleDistributions = np.divide(femaleQuestionResponses, len(femaleResponses))
    maleDistributions = np.divide(maleQuestionResponses, len(maleResponses))

    return femaleDistributions, maleDistributions

def drawBarGraph(maleDist, femaleDist, questionTitle):
    nGroups = 5
    fig, p = plt.subplots()
    index = np.arange(nGroups)
    barWidth = 0.35

    maleBar = p.bar(index, tuple(maleDist), barWidth, color='g', label='Men')
    femaleBar = p.bar(index + barWidth, tuple(femaleDist), barWidth, color='b', label='Female')

    p.set_xlabel('Responses')
    p.set_ylabel('Distributions')
    p.set_title("{0} Distributions for Men and Women".format(questionTitle))
    p.set_xticks(index + barWidth / 2)
    p.set_xticklabels(('1', '2', '3', '4', '5'))
    p.legend()

    fig.tight_layout()
    plt.show()

def calculateHellbringerDistances(maleDist, femaleDist, questionTitle):
    hellinDict = {}
    divSqRtTwo = 1 / math.sqrt(2)

    tmp = math.sqrt(np.sum(np.square(np.subtract(np.sqrt(maleDist), np.sqrt(femaleDist)))))
    hellin = tmp * divSqRtTwo
    return hellin

        

for i in range(26):
    colLetter = chr(65 + i)            
    questionName = "Q%d" % (i + 1)

    femaleProb, maleProb = distributionOfColumns(colLetter)
    #drawBarGraph(femaleProb, maleProb, questionName)
    hellinDist = calculateHellbringerDistances(maleProb, femaleProb, questionName)
    maleQuestionDistributions.update({questionName:maleProb})
    femaleQuestionDistributions.update({questionName:femaleProb})
    hellingerDistances.update({questionName:hellinDist})